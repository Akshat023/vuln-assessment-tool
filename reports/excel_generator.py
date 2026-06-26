"""
Excel Report Generator
======================
Generates a professional vulnerability assessment Excel workbook
from enriched scan findings using OpenPyXL.

Workbook structure:
    Sheet 1: Executive Summary  — scan metadata, AI summary, severity counts
    Sheet 2: Findings Table     — all vulnerabilities, color-coded by severity
    Sheet 3: Remediation Checklist — one row per finding, checkbox to track fixes

Install:
    pip install openpyxl

Usage:
    from reports.excel_generator import ExcelGenerator
    generator = ExcelGenerator()
    path = generator.generate(scan_data, output_path="report.xlsx")
"""

import os
import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Color palette
# ──────────────────────────────────────────────
COLORS = {
    "critical_bg":   "7B0000",
    "critical_text": "FFFFFF",
    "high_bg":       "C0392B",
    "high_text":     "FFFFFF",
    "medium_bg":     "E67E22",
    "medium_text":   "FFFFFF",
    "low_bg":        "2980B9",
    "low_text":      "FFFFFF",
    "info_bg":       "7F8C8D",
    "info_text":     "FFFFFF",
    "header_bg":     "1a1a2e",
    "header_text":   "FFFFFF",
    "subheader_bg":  "2C3E6E",
    "subheader_text":"A8C6FA",
    "alt_row":       "F8F9FA",
    "white":         "FFFFFF",
    "border":        "BDC3C7",
    "done_bg":       "D5F5E3",
    "pending_bg":    "FDECEA",
}

SEVERITY_COLORS = {
    "Critical":      ("7B0000", "FFE5E5"),
    "High":          ("C0392B", "FDECEA"),
    "Medium":        ("E67E22", "FEF3E2"),
    "Low":           ("2980B9", "EBF5FB"),
    "Informational": ("7F8C8D", "F2F3F4"),
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _make_border(color: str = "BDC3C7") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _make_font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


class ExcelGenerator:
    """
    Generates Excel vulnerability reports from scan data.

    Args:
        output_dir: Directory to save generated Excel files
    """

    def __init__(self, output_dir: str = "reports/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ──────────────────────────────────────────
    # Sheet 1: Executive Summary
    # ──────────────────────────────────────────

    def _build_summary_sheet(self, ws, scan_data: dict):
        """Build the Executive Summary sheet."""

        scan_id      = scan_data.get("scan_id", "N/A")
        target_url   = scan_data.get("url") or scan_data.get("target", "N/A")
        created_at   = scan_data.get("created_at", datetime.utcnow().isoformat())
        summary      = scan_data.get("summary", {})
        exec_summary = scan_data.get("executive_summary", "No executive summary available.")

        try:
            date_str = datetime.fromisoformat(created_at).strftime("%B %d, %Y %H:%M UTC")
        except Exception:
            date_str = created_at

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        # ── Title ──────────────────────────────
        ws.merge_cells("A1:D1")
        ws["A1"] = "AI-Powered Vulnerability Assessment Report"
        ws["A1"].font = _make_font(bold=True, color=COLORS["header_text"], size=16)
        ws["A1"].fill = _make_fill(COLORS["header_bg"])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:D2")
        ws["A2"] = "Powered by OWASP ZAP + Groq AI (Llama 3.3 70B)"
        ws["A2"].font = _make_font(color=COLORS["subheader_text"], size=11)
        ws["A2"].fill = _make_fill(COLORS["subheader_bg"])
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # ── Scan metadata ──────────────────────
        ws.merge_cells("A4:D4")
        ws["A4"] = "Scan Information"
        ws["A4"].font = _make_font(bold=True, color=COLORS["header_text"], size=12)
        ws["A4"].fill = _make_fill(COLORS["subheader_bg"])
        ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[4].height = 24

        meta_rows = [
            ("Target URL",   target_url),
            ("Scan Date",    date_str),
            ("Scan ID",      scan_id),
            ("Total Findings", str(summary.get("total", 0))),
        ]

        for i, (label, value) in enumerate(meta_rows, start=5):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = _make_font(bold=True, size=11)
            ws[f"A{i}"].fill = _make_fill(COLORS["alt_row"])
            ws[f"A{i}"].border = _make_border()

            ws.merge_cells(f"B{i}:D{i}")
            ws[f"B{i}"] = value
            ws[f"B{i}"].font = _make_font(size=11)
            ws[f"B{i}"].border = _make_border()
            ws[f"B{i}"].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[i].height = 20

        # ── Severity summary boxes ─────────────
        row = 11
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Severity Summary"
        ws[f"A{row}"].font = _make_font(bold=True, color=COLORS["header_text"], size=12)
        ws[f"A{row}"].fill = _make_fill(COLORS["subheader_bg"])
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24

        row += 1
        severity_cols = [
            ("Critical", "A", "7B0000", "FFE5E5"),
            ("High",     "B", "C0392B", "FDECEA"),
            ("Medium",   "C", "E67E22", "FEF3E2"),
            ("Low",      "D", "2980B9", "EBF5FB"),
        ]

        for sev, col, text_color, bg_color in severity_cols:
            count = summary.get(sev.lower(), 0)
            cell_label = ws[f"{col}{row}"]
            cell_label.value = sev
            cell_label.font = _make_font(bold=True, color=text_color, size=11)
            cell_label.fill = _make_fill(bg_color)
            cell_label.alignment = Alignment(horizontal="center")
            cell_label.border = _make_border(text_color)

            cell_count = ws[f"{col}{row + 1}"]
            cell_count.value = count
            cell_count.font = _make_font(bold=True, color=text_color, size=20)
            cell_count.fill = _make_fill(bg_color)
            cell_count.alignment = Alignment(horizontal="center", vertical="center")
            cell_count.border = _make_border(text_color)

        ws.row_dimensions[row].height = 20
        ws.row_dimensions[row + 1].height = 36

        # ── Executive Summary ──────────────────
        row += 4
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Executive Summary (AI-Generated)"
        ws[f"A{row}"].font = _make_font(bold=True, color=COLORS["header_text"], size=12)
        ws[f"A{row}"].fill = _make_fill(COLORS["subheader_bg"])
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24

        row += 1
        ws.merge_cells(f"A{row}:D{row + 6}")
        ws[f"A{row}"] = exec_summary
        ws[f"A{row}"].font = _make_font(size=11)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"A{row}"].border = _make_border()
        ws.row_dimensions[row].height = 120

        # ── Disclaimer ─────────────────────────
        row += 9
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "⚠ Disclaimer: This report is generated by an automated tool and does not replace a professional penetration test."
        ws[f"A{row}"].font = _make_font(color="888888", size=9)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    # ──────────────────────────────────────────
    # Sheet 2: Findings Table
    # ──────────────────────────────────────────

    def _build_findings_sheet(self, ws, findings: list):
        """Build the full findings table sheet."""

        # Column widths
        col_widths = [12, 28, 30, 8, 40, 45, 45]
        col_names  = ["Severity", "Vulnerability", "OWASP Category", "CVSS",
                      "Affected URL", "Recommendation", "Business Impact"]

        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Header row
        for col, name in enumerate(col_names, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = _make_font(bold=True, color=COLORS["header_text"], size=11)
            cell.fill = _make_fill(COLORS["header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _make_border()
        ws.row_dimensions[1].height = 28

        # Freeze header row
        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, finding in enumerate(findings, start=2):
            sev = finding.get("severity", "Low")
            text_color, bg_color = SEVERITY_COLORS.get(sev, ("000000", "FFFFFF"))
            row_bg = bg_color if row_idx % 2 == 0 else COLORS["white"]

            values = [
                sev,
                finding.get("vuln_type", ""),
                finding.get("owasp_category", ""),
                finding.get("cvss_score", ""),
                finding.get("affected_url", ""),
                finding.get("recommendation", finding.get("solution", "")),
                finding.get("business_impact", ""),
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _make_border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Severity cell gets colored badge styling
                if col_idx == 1:
                    cell.font = _make_font(bold=True, color=text_color, size=10)
                    cell.fill = _make_fill(bg_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # CVSS score centered and colored
                elif col_idx == 4:
                    cell.font = _make_font(bold=True, color=text_color, size=11)
                    cell.fill = _make_fill(bg_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = _make_font(size=10)
                    cell.fill = _make_fill(row_bg)

            ws.row_dimensions[row_idx].height = 60

        # Auto-filter on header
        ws.auto_filter.ref = f"A1:G{len(findings) + 1}"

    # ──────────────────────────────────────────
    # Sheet 3: Remediation Checklist
    # ──────────────────────────────────────────

    def _build_checklist_sheet(self, ws, findings: list):
        """Build the remediation checklist sheet."""

        col_widths = [10, 12, 28, 40, 50, 12]
        col_names  = ["✓ Fixed", "Severity", "Vulnerability", "Affected URL",
                      "Recommended Fix", "CVSS"]

        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "Remediation Checklist — Track your fixes here"
        ws["A1"].font = _make_font(bold=True, color=COLORS["header_text"], size=13)
        ws["A1"].fill = _make_fill(COLORS["header_bg"])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Header row
        for col, name in enumerate(col_names, 1):
            cell = ws.cell(row=2, column=col, value=name)
            cell.font = _make_font(bold=True, color=COLORS["header_text"], size=11)
            cell.fill = _make_fill(COLORS["subheader_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _make_border()
        ws.row_dimensions[2].height = 24

        ws.freeze_panes = "A3"

        # Only include Medium+ for checklist (actionable items)
        priority_findings = [f for f in findings if f.get("severity") in ("Critical", "High", "Medium")]
        if not priority_findings:
            priority_findings = findings  # fallback: include all

        for row_idx, finding in enumerate(priority_findings, start=3):
            sev = finding.get("severity", "Low")
            text_color, bg_color = SEVERITY_COLORS.get(sev, ("000000", "FFFFFF"))

            # Checkbox column — user fills this in
            check_cell = ws.cell(row=row_idx, column=1, value="☐")
            check_cell.font = _make_font(size=14, color="666666")
            check_cell.fill = _make_fill(COLORS["pending_bg"])
            check_cell.alignment = Alignment(horizontal="center", vertical="center")
            check_cell.border = _make_border()

            values = [
                sev,
                finding.get("vuln_type", ""),
                finding.get("affected_url", ""),
                finding.get("recommendation", finding.get("solution", ""))[:300],
                finding.get("cvss_score", ""),
            ]

            for col_idx, value in enumerate(values, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _make_border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if col_idx == 2:  # Severity
                    cell.font = _make_font(bold=True, color=text_color, size=10)
                    cell.fill = _make_fill(bg_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 6:  # CVSS
                    cell.font = _make_font(bold=True, color=text_color, size=11)
                    cell.fill = _make_fill(bg_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = _make_font(size=10)
                    cell.fill = _make_fill(
                        COLORS["alt_row"] if row_idx % 2 == 0 else COLORS["white"]
                    )

            ws.row_dimensions[row_idx].height = 55

        ws.auto_filter.ref = f"A2:F{len(priority_findings) + 2}"

    # ──────────────────────────────────────────
    # Public API: generate
    # ──────────────────────────────────────────

    def generate(self, scan_data: dict, output_filename: str = None) -> str:
        """
        Generate an Excel report from scan data.

        Args:
            scan_data:        Full scan result dict
            output_filename:  Optional custom filename

        Returns:
            Path to the generated Excel file
        """
        scan_id = scan_data.get("scan_id", "unknown")
        findings = scan_data.get("findings", [])

        if output_filename is None:
            output_filename = f"report_{scan_id}.xlsx"

        output_path = self.output_dir / output_filename

        logger.info(f"Generating Excel report for scan {scan_id} with {len(findings)} findings")

        wb = Workbook()

        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        self._build_summary_sheet(ws1, scan_data)

        # Sheet 2: All Findings
        ws2 = wb.create_sheet("Findings")
        self._build_findings_sheet(ws2, findings)

        # Sheet 3: Remediation Checklist
        ws3 = wb.create_sheet("Remediation Checklist")
        self._build_checklist_sheet(ws3, findings)

        wb.save(str(output_path))
        logger.info(f"Excel saved to: {output_path}")
        return str(output_path)


# ──────────────────────────────────────────────
# Quick test
# ──────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    for fname in ["scan_result_enriched.json", "scan_result.json"]:
        if os.path.exists(fname):
            with open(fname) as f:
                raw = json.load(f)
            print(f"Loaded from {fname}")
            break
    else:
        print("No scan data found. Run zap_scanner.py first.")
        sys.exit(1)

    # Patch missing fields
    if not raw.get("scan_id") and raw.get("findings"):
        raw["scan_id"] = raw["findings"][0].get("scan_id", "unknown")
    if not raw.get("url") and raw.get("target"):
        raw["url"] = raw["target"]
    if "summary" not in raw:
        findings = raw.get("findings", [])
        summary = {"total": len(findings), "critical": 0, "high": 0, "medium": 0, "low": 0}
        for f in findings:
            sev = f.get("severity", "Low").lower()
            if sev in summary:
                summary[sev] += 1
        raw["summary"] = summary

    generator = ExcelGenerator(output_dir="reports/output")
    path = generator.generate(raw)
    print(f"\nExcel report generated: {path}")
    print("Open it to see all 3 sheets!")